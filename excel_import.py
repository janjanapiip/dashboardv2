"""Parse the SPP master workbook.

Supports two layouts:

  OLD (single row per lab, KETERANGAN column with dd/mm parseable text):
    Row 12 = Lab #1 (FR/JLH/DRS per date + free-text KETERANGAN at end)
    Row 13 = Lab #2
    ...

  NEW (2 rows per lab, Kegiatan row directly under data row):
    Row 12 = Lab #1 data (FR/JLH/DRS per date)
    Row 13 = Lab #1 "Kegiatan" — one merged 3-col cell per date with activity text
    Row 14 = Lab #2 data
    Row 15 = Lab #2 Kegiatan
    ...
    Col B on Kegiatan rows literally reads "Kegiatan" — that's the detection marker.

  Both layouts produce the same output dicts (rows / keterangan / details), so the
  downstream importer and app don't need to know which format they came from.
"""
import re
from calendar import monthrange
from difflib import SequenceMatcher
import openpyxl
from labs import LABS, MONTHS_LOOKUP

FIRST_DATA_ROW = 12

LAB_NAME_BY_ID = {lab_id: name for lab_id, _, name in LABS}

# Backwards-compat: parses "05/01 : PUKP, Pasis : Try out" style lines in OLD KETERANGAN cells.
_DETAIL_LINE_RE = re.compile(
    r"^\s*(?P<day>\d{1,2})\s*/\s*(?P<month>\d{1,2})(?:\s*/\s*(?P<year>\d{2,4}))?\s*/?\s*[:\-]\s*(?P<body>.+?)\s*$"
)


def parse_detail_lines(text, sheet_year: int, sheet_month: int):
    """OLD-format helper: split a KETERANGAN cell into per-date entries + leftover free text."""
    if text is None:
        return [], []
    details: list[dict] = []
    leftover: list[str] = []
    days_in_month = monthrange(sheet_year, sheet_month)[1]

    for raw in str(text).splitlines():
        line = raw.strip()
        if not line:
            continue
        m = _DETAIL_LINE_RE.match(line)
        if not m:
            leftover.append(line)
            continue
        try:
            day = int(m.group("day"))
            month = int(m.group("month"))
        except (TypeError, ValueError):
            leftover.append(line)
            continue
        if month != sheet_month or day < 1 or day > days_in_month:
            leftover.append(line)
            continue
        body = m.group("body").strip()
        if ":" in body:
            users, activity = body.split(":", 1)
        else:
            users, activity = "", body
        details.append({
            "day": day,
            "users": users.strip(),
            "activity": activity.strip(),
        })
    return details, leftover


def _split_users_activity(text: str):
    """New-format Kegiatan cell may still use 'users : activity' — try to split."""
    if not text:
        return "", ""
    s = str(text).strip()
    if ":" in s:
        users, activity = s.split(":", 1)
        return users.strip(), activity.strip()
    return "", s


def _normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _find_sheet(wb):
    for name in wb.sheetnames:
        if "master" in name.lower() or "utilisas" in name.lower() or "rekap" in name.lower():
            return wb[name]
    return wb[wb.sheetnames[0]]


def _parse_period(ws):
    year = None
    month = None
    for r in range(1, 12):
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if not v:
                continue
            s = str(v).lower()
            m_year = re.search(r"(20\d{2})", s)
            if m_year:
                year = int(m_year.group(1))
            for name_low, num in MONTHS_LOOKUP.items():
                if name_low in s:
                    month = num
                    break
            if year and month:
                return year, month
    return year, month


def _is_new_format(ws) -> bool:
    """Detect NEW layout by checking if row 13 col B contains 'Kegiatan'."""
    v = ws.cell(row=13, column=2).value
    return bool(v) and _normalize(str(v)) == "kegiatan"


def _build_row_to_lab_map(ws, stride: int) -> dict[int, int]:
    """Walk the sheet's actual lab rows and assign each to the best-matching lab_id
    from labs.LABS. Direction is IMPORTANT: iterating labs → best-row (the old logic)
    causes retired labs to steal rows from active labs when the sheet only contains
    the 8 non-retired labs. Iterating rows → best-lab makes retired labs unmatched.

    Threshold 0.75 is deliberately high so unrelated names don't match. Positional
    fallback is deliberately removed here — if the sheet's row name is unreadable,
    we skip it and log rather than mis-assigning."""
    # Scan up to ~13 possible lab rows (with stride) to be safe if sheet later grows.
    max_rows = 13 * max(stride, 1) + FIRST_DATA_ROW
    result: dict[int, int] = {}
    used_labs: set[int] = set()
    for r in range(FIRST_DATA_ROW, max_rows, stride):
        raw = ws.cell(row=r, column=2).value
        if not raw:
            continue
        target = _normalize(str(raw))
        if not target:
            continue
        best_ratio, best_lab_id = 0.0, None
        for lab_id, _, lab_name in LABS:
            if lab_id in used_labs:
                continue
            ratio = SequenceMatcher(None, _normalize(lab_name), target).ratio()
            if ratio > best_ratio:
                best_ratio, best_lab_id = ratio, lab_id
        if best_lab_id is not None and best_ratio >= 0.75:
            result[r] = best_lab_id
            used_labs.add(best_lab_id)
    return result


def _read_kegiatan_new(ws, keg_row: int, base_col: int) -> str:
    """New format: read the Kegiatan text at (keg_row, base_col). May be part of a merged
    range whose anchor is base_col; if empty at base_col but the range spans, the merged
    anchor holds the value. openpyxl returns the anchor value at any cell of the range."""
    v = ws.cell(row=keg_row, column=base_col).value
    if v is None:
        # Sometimes the value is stored on the anchor cell of a merged range.
        # For safety, scan the 3 subcolumns.
        for off in (1, 2):
            v = ws.cell(row=keg_row, column=base_col + off).value
            if v is not None:
                break
    return str(v).strip() if v is not None else ""


def _parse_sheet(ws):
    year, month = _parse_period(ws)
    if not year or not month:
        raise ValueError(
            "Tidak ada keterangan bulan/tahun pada baris 6-8 ('BULAN : <Bulan> <Tahun>')."
        )

    days_in_month = monthrange(year, month)[1]
    is_new = _is_new_format(ws)
    stride = 2 if is_new else 1

    # Build sheet-row → lab_id map by walking actual rows in the sheet.
    # Labs not present in the sheet (e.g. retired ones) simply don't appear here.
    row_to_lab = _build_row_to_lab_map(ws, stride)

    rows = []
    keterangan = {}
    details = []

    for data_row, lab_id in row_to_lab.items():
        keg_row = data_row + 1 if is_new else None

        for d in range(1, days_in_month + 1):
            base_col = 6 + (d - 1) * 3  # F=6 -> day 1
            fr_v = ws.cell(row=data_row, column=base_col).value
            jlh_v = ws.cell(row=data_row, column=base_col + 1).value
            drs_v = ws.cell(row=data_row, column=base_col + 2).value
            fr = int(fr_v) if isinstance(fr_v, (int, float)) and fr_v is not None else 0
            jlh = int(jlh_v) if isinstance(jlh_v, (int, float)) and jlh_v is not None else 0
            try:
                drs = float(drs_v) if drs_v is not None and str(drs_v).strip() != "" else 0.0
            except (TypeError, ValueError):
                drs = 0.0
            if fr or jlh or drs:
                rows.append({"lab_id": lab_id, "day": d, "fr": fr, "jlh": jlh, "drs": drs})

            # NEW format: pull Kegiatan text for this specific date from keg_row
            if is_new:
                keg_text = _read_kegiatan_new(ws, keg_row, base_col)
                if keg_text:
                    users, activity = _split_users_activity(keg_text)
                    details.append({
                        "lab_id": lab_id,
                        "day": d,
                        "users": users,
                        "activity": activity,
                    })

        # KETERANGAN column at the end — used in both layouts as a free-text note.
        ket_col = 6 + days_in_month * 3
        ket = ws.cell(row=data_row, column=ket_col).value
        if ket and str(ket).strip():
            if is_new:
                # Whole-month note only; no dd/mm parsing needed.
                keterangan[lab_id] = str(ket).strip()
            else:
                parsed_details, leftover = parse_detail_lines(ket, year, month)
                for item in parsed_details:
                    details.append({"lab_id": lab_id, **item})
                residual = "\n".join(leftover).strip()
                if residual:
                    keterangan[lab_id] = residual

    return {
        "year": year, "month": month,
        "rows": rows, "keterangan": keterangan,
        "details": details,
    }


def parse_master_format(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_sheet(wb)
    return _parse_sheet(ws)


def parse_all_sheets(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    results = []
    for name in wb.sheetnames:
        ws = wb[name]
        clean_name = name.strip()
        try:
            parsed = _parse_sheet(ws)
            results.append({
                "sheet": clean_name,
                "ok": True,
                "year": parsed["year"],
                "month": parsed["month"],
                "rows": parsed["rows"],
                "keterangan": parsed["keterangan"],
                "details": parsed["details"],
            })
        except Exception as e:
            results.append({"sheet": clean_name, "ok": False, "error": str(e)})
    return results
