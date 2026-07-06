"""Build REKAP UTILISASI SPP v4 tabel Mei-Des.xlsx — 2-row-per-lab layout.

New layout (per lab, per date):
  Row N   (data):    | FR | JLH | DRS |  for each day
  Row N+1 (Kegiatan): merged 3-col cell holding activity description

Formulas:
  JF     = SUM of all FR columns for this lab's data row
  JP     = SUM of all JLH columns for this lab's data row
  DURASI = SUM of all DRS columns for this lab's data row

Preserves existing JUNI data from `master/juni_extracted.json` (if present).
Other months (MEI, JULI..DES) are built empty.

Days per month (2026): MEI=31, JUNI=30, JULI=31, AGT=31, SEP=30, OKT=31, NOV=30, DES=31.

Uses the JUNI sheet in the current master file as the STYLE source (fonts, borders,
column widths, row heights, header block). Data rows are then rebuilt from scratch.
"""
from __future__ import annotations
import json
from calendar import monthrange
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
MASTER_DIR = HERE.parent / "master"
SRC = MASTER_DIR / "REKAP UTILISASI SPP v3 tabel Mei-Des.xlsx"
DST = MASTER_DIR / "REKAP UTILISASI SPP v3 tabel Mei-Des.xlsx"  # overwrite in place
JUNI_JSON = MASTER_DIR / "juni_extracted.json"

YEAR = 2026
MONTHS_ID = {
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER",
}

# 8 labs (post-v3 removal), in canonical order.
LABS = [
    ("CHL",  "CARGO HANDLING LABORATORY (CHL)"),
    ("EEL",  "ELECTRIC AND ELECTRONIC LABORATORY (EEL)"),
    ("EWS",  "ENGINEERING WORKSHOP"),
    ("MEL",  "MARINE ENGINEERING LABORATORY (MEL)"),
    ("CBT",  "COMPUTER BASED TRAINING 1 & 2 (CBT)"),
    ("ERCS", "ENGINE ROOM CERTIFICATION SIMULATOR (ERCS)"),
    ("LTL",  "LANGUAGE TRAINING LABORATORY (LTL)"),
    ("ACSL", "AUTOMATIC CONTROL SYSTEM LABORATORY (ACSL)"),
]

# Layout constants
HEADER_ROWS = 11         # rows 1..11 are static header (title, BULAN, column headers, day numbers, FR/JLH/DRS labels)
FIRST_DATA_ROW = 12      # Lab 1 data row
ROWS_PER_LAB = 2         # data row + Kegiatan row
DATE_START_COL = 6       # col F = day 1's FR column

# Styles
thin = Side(border_style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
KEG_FILL = PatternFill("solid", fgColor="FFF2CC")


def date_col(day: int) -> int:
    """day 1 -> col F (6), day 2 -> col I (9), ..."""
    return DATE_START_COL + (day - 1) * 3


def build_sum_formula(lab_data_row: int, offset: int, days: int) -> str:
    """SUM formula referencing FR/JLH/DRS column at each day.
    offset: 0=FR, 1=JLH, 2=DRS."""
    cells = [f"{get_column_letter(date_col(d) + offset)}{lab_data_row}" for d in range(1, days + 1)]
    return "=SUM(" + ",".join(cells) + ")"


def build_sheet(wb, month_num: int, month_name: str, existing_data: list[dict] | None):
    """Build one month sheet in the new 2-row-per-lab layout."""
    if month_name in wb.sheetnames:
        del wb[month_name]
    ws = wb.create_sheet(title=month_name)

    days = monthrange(YEAR, month_num)[1]

    # ---- Header block (rows 1-8): institutional identity ----
    ws.cell(row=1, column=3, value="KEMENTERIAN PERHUBUNGAN")
    ws.cell(row=2, column=3, value="BADAN PENGEMBANGAN SUMBER DAYA MANUSIA PERHUBUNGAN")
    ws.cell(row=3, column=3, value="SEKOLAH TINGGI ILMU PELAYARAN")
    ws.cell(row=4, column=3, value="UNIT SARANA PRAKTEK PELAUT")
    for r in (1, 2, 3, 4):
        ws.cell(row=r, column=3).font = BOLD
        ws.cell(row=r, column=3).alignment = CENTER

    ws.cell(row=6, column=2, value="DOKUMEN")
    ws.cell(row=6, column=3, value=f": Rekapitulasi Utilisasi Sarana Praktek Pelaut")
    ws.cell(row=7, column=2, value="BULAN")
    ws.cell(row=7, column=3, value=f": {month_name.title()} {YEAR}")
    ws.cell(row=6, column=2).font = BOLD
    ws.cell(row=7, column=2).font = BOLD

    # ---- Row 9-11: table header ----
    # Row 9: NO / NAMA LAB / JF / JP / DURASI (JAM) / TANGGAL KERJA
    hdr_labels = [(1, "NO"), (2, "NAMA LAB"), (3, "JF"), (4, "JP"), (5, "DURASI (JAM)")]
    for col, txt in hdr_labels:
        c = ws.cell(row=9, column=col, value=txt)
        c.font = BOLD
        c.alignment = CENTER
        c.fill = HEADER_FILL
        c.border = BORDER
    # TANGGAL KERJA spans all date columns on row 9
    last_date_col = date_col(days) + 2
    keg_col = last_date_col + 1
    c = ws.cell(row=9, column=DATE_START_COL, value="TANGGAL KERJA")
    c.font = BOLD; c.alignment = CENTER; c.fill = HEADER_FILL; c.border = BORDER
    ws.merge_cells(start_row=9, start_column=DATE_START_COL, end_row=9, end_column=last_date_col)
    # KETERANGAN column (last, kept for backward compat / free-text notes)
    c = ws.cell(row=9, column=keg_col, value="KETERANGAN")
    c.font = BOLD; c.alignment = CENTER; c.fill = HEADER_FILL; c.border = BORDER

    # Merge the fixed columns down rows 9..11
    for col in (1, 2, 3, 4, 5):
        ws.merge_cells(start_row=9, start_column=col, end_row=11, end_column=col)
    ws.merge_cells(start_row=9, start_column=keg_col, end_row=11, end_column=keg_col)

    # Row 10: date numbers, each spans 3 columns (FR/JLH/DRS)
    for d in range(1, days + 1):
        c0 = date_col(d)
        cell = ws.cell(row=10, column=c0, value=d)
        cell.font = BOLD; cell.alignment = CENTER; cell.fill = HEADER_FILL; cell.border = BORDER
        ws.merge_cells(start_row=10, start_column=c0, end_row=10, end_column=c0 + 2)

    # Row 11: FR / JLH / DRS labels
    for d in range(1, days + 1):
        c0 = date_col(d)
        for i, lbl in enumerate(("FR", "JLH", "DRS")):
            cell = ws.cell(row=11, column=c0 + i, value=lbl)
            cell.font = BOLD; cell.alignment = CENTER; cell.fill = HEADER_FILL; cell.border = BORDER

    # ---- Data rows: 2 per lab ----
    existing_by_code = {}
    if existing_data:
        for lab in existing_data:
            # JSON serializes dict keys as strings — coerce to int here so day lookups work
            days_int = {int(k): v for k, v in lab["days"].items()}
            existing_by_code[lab["code"]] = {**lab, "days": days_int}

    for idx, (code, name) in enumerate(LABS):
        data_row = FIRST_DATA_ROW + idx * ROWS_PER_LAB
        keg_row = data_row + 1

        # NO
        ws.cell(row=data_row, column=1, value=idx + 1).alignment = CENTER
        ws.cell(row=data_row, column=1).border = BORDER
        # NAMA LAB
        ws.cell(row=data_row, column=2, value=name).alignment = LEFT_WRAP
        ws.cell(row=data_row, column=2).border = BORDER
        # Row (data_row+1) col B = "Kegiatan"
        kc = ws.cell(row=keg_row, column=2, value="Kegiatan")
        kc.alignment = CENTER
        kc.font = Font(italic=True, color="808080", size=9)
        kc.border = BORDER
        kc.fill = KEG_FILL

        # Merge NO and NAMA LAB across data + Kegiatan rows
        ws.merge_cells(start_row=data_row, start_column=1, end_row=keg_row, end_column=1)
        # NAMA LAB stays only on data_row; keg row col B holds the "Kegiatan" tag

        # JF/JP/DURASI formulas
        ws.cell(row=data_row, column=3, value=build_sum_formula(data_row, 0, days))
        ws.cell(row=data_row, column=4, value=build_sum_formula(data_row, 1, days))
        ws.cell(row=data_row, column=5, value=build_sum_formula(data_row, 2, days))
        for col in (3, 4, 5):
            cc = ws.cell(row=data_row, column=col)
            cc.alignment = CENTER; cc.border = BORDER; cc.font = BOLD
        # Merge JF/JP/DURASI across data + keg rows
        for col in (3, 4, 5):
            ws.merge_cells(start_row=data_row, start_column=col, end_row=keg_row, end_column=col)

        # Populate existing data (JUNI only) + border every FR/JLH/DRS cell
        existing = existing_by_code.get(code)
        for d in range(1, days + 1):
            c0 = date_col(d)
            for i in range(3):
                cell = ws.cell(row=data_row, column=c0 + i)
                cell.alignment = CENTER
                cell.border = BORDER
            if existing and d in existing["days"]:
                dd = existing["days"][d]
                if dd["fr"]:
                    ws.cell(row=data_row, column=c0).value = dd["fr"]
                if dd["jlh"]:
                    ws.cell(row=data_row, column=c0 + 1).value = dd["jlh"]
                if dd["drs"]:
                    ws.cell(row=data_row, column=c0 + 2).value = dd["drs"]

            # Kegiatan row: merge 3 cols per date, style, populate
            ws.merge_cells(start_row=keg_row, start_column=c0, end_row=keg_row, end_column=c0 + 2)
            kcell = ws.cell(row=keg_row, column=c0)
            kcell.alignment = LEFT_WRAP
            kcell.border = BORDER
            kcell.fill = KEG_FILL
            kcell.font = Font(size=9)
            if existing and d in existing["days"] and existing["days"][d].get("kegiatan"):
                kcell.value = existing["days"][d]["kegiatan"]

        # KETERANGAN (last col) — free-text, merged across data+keg
        kt = ws.cell(row=data_row, column=keg_col)
        kt.alignment = LEFT_WRAP
        kt.border = BORDER
        ws.merge_cells(start_row=data_row, start_column=keg_col, end_row=keg_row, end_column=keg_col)

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 42
    for col in (3, 4, 5):
        ws.column_dimensions[get_column_letter(col)].width = 8
    for d in range(1, days + 1):
        c0 = date_col(d)
        for i in range(3):
            ws.column_dimensions[get_column_letter(c0 + i)].width = 6
    ws.column_dimensions[get_column_letter(keg_col)].width = 30

    # ---- Row heights ----
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 16
    for idx in range(len(LABS)):
        data_row = FIRST_DATA_ROW + idx * ROWS_PER_LAB
        keg_row = data_row + 1
        ws.row_dimensions[data_row].height = 22
        ws.row_dimensions[keg_row].height = 30  # taller for wrapped text

    # Freeze the header + label columns
    ws.freeze_panes = ws.cell(row=12, column=DATE_START_COL)


def build():
    # Load existing workbook and preserve any styles we want; we're rebuilding sheets from scratch
    if SRC.exists():
        wb = load_workbook(SRC)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        # remove default sheet
        for s in list(wb.sheetnames):
            del wb[s]

    # Load JUNI data if available
    juni_data = None
    if JUNI_JSON.exists():
        with open(JUNI_JSON, encoding="utf-8") as f:
            juni_data = json.load(f)
        print(f"Loaded JUNI data: {sum(len(l['days']) for l in juni_data)} filled days")

    # Build each month
    for month_num, month_name in MONTHS_ID.items():
        data = juni_data if month_num == 6 else None
        build_sheet(wb, month_num, month_name, data)
        print(f"  Built {month_name} ({monthrange(YEAR, month_num)[1]} days)")

    # Ensure sheet ordering: MEI..DESEMBER
    ordered = [MONTHS_ID[m] for m in sorted(MONTHS_ID.keys())]
    # Drop any leftover sheets not in our target set
    for name in list(wb.sheetnames):
        if name not in ordered:
            del wb[name]
    # Reorder
    wb._sheets = [wb[n] for n in ordered]

    wb.save(DST)
    print(f"\nWrote {DST}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    build()
