"""Add lab #9 'Lain-lain (rooftop lantai 3, selasar, dll)' to all sheets
in the master v3 workbook.

Layout (NEW format = 2 rows per lab):
    Row 26 = lab #8 (ACSL) data
    Row 27 = lab #8 Kegiatan
    Row 28 = lab #9 (LAI)  data  ← we add
    Row 29 = lab #9 Kegiatan     ← we add

Copies style/border/fill/font/alignment/merges from rows 26–27 into 28–29
and rewrites SUM formulas so they reference the correct row.
"""
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MASTER = Path(__file__).resolve().parents[1] / "master" / "REKAP UTILISASI SPP v3 tabel Mei-Des.xlsx"

# Column letters that carry per-date sub-columns (F..CQ, step 3)
DATE_BASE_COLS = list(range(6, 6 + 31 * 3, 3))  # F=6 for day 1, +3 each

SRC_DATA_ROW = 26  # lab #8 data row (last existing) → we copy from here
SRC_KEG_ROW = 27
DST_DATA_ROW = 28  # lab #9 data
DST_KEG_ROW = 29

LAB_NUMBER = 9
LAB_NAME = "LAIN-LAIN (ROOFTOP LANTAI 3, SELASAR, DLL)"

def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

def rewrite_formula(formula: str, src_row: int, dst_row: int) -> str:
    """Replace ONLY row numbers matching src_row (e.g. 26) with dst_row (28).
    Naïve textual replacement is fine here because the SUM formulas reference
    only cells on the same row (row 26 → row 28)."""
    if not formula or not formula.startswith("="):
        return formula
    # Replace `26)` `26,` `26$` etc — use word boundary before the number
    import re
    return re.sub(rf'(?<=[A-Za-z]){src_row}\b', str(dst_row), formula)

def process_sheet(ws) -> str:
    """Process one sheet; return status message."""
    # Idempotency: skip if row 28 col B already has the LAI name
    existing = ws.cell(row=DST_DATA_ROW, column=2).value
    if existing and "lain" in str(existing).lower():
        return f"skip (row 28 already has '{existing}')"

    # 1) Copy row heights
    if SRC_DATA_ROW in ws.row_dimensions:
        ws.row_dimensions[DST_DATA_ROW].height = ws.row_dimensions[SRC_DATA_ROW].height
    if SRC_KEG_ROW in ws.row_dimensions:
        ws.row_dimensions[DST_KEG_ROW].height = ws.row_dimensions[SRC_KEG_ROW].height

    # 2) Copy every cell in both rows (style + value/formula)
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        # Data row 26 → 28
        src = ws.cell(row=SRC_DATA_ROW, column=col)
        dst = ws.cell(row=DST_DATA_ROW, column=col)
        copy_cell_style(src, dst)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = rewrite_formula(src.value, SRC_DATA_ROW, DST_DATA_ROW)
        else:
            dst.value = None  # leave data cells empty
        # Kegiatan row 27 → 29
        src_k = ws.cell(row=SRC_KEG_ROW, column=col)
        dst_k = ws.cell(row=DST_KEG_ROW, column=col)
        copy_cell_style(src_k, dst_k)
        if isinstance(src_k.value, str) and src_k.value.startswith("="):
            dst_k.value = rewrite_formula(src_k.value, SRC_KEG_ROW, DST_KEG_ROW)
        else:
            dst_k.value = None

    # 3) Set the identity cells
    ws.cell(row=DST_DATA_ROW, column=1).value = LAB_NUMBER  # NO
    ws.cell(row=DST_DATA_ROW, column=2).value = LAB_NAME    # NAMA LAB
    ws.cell(row=DST_KEG_ROW, column=2).value = "Kegiatan"

    # 4) Copy merged ranges from src rows → dst rows
    src_merges = [str(m) for m in ws.merged_cells.ranges
                  if m.min_row in (SRC_DATA_ROW, SRC_KEG_ROW)
                  and m.max_row in (SRC_DATA_ROW, SRC_KEG_ROW)]
    for m_str in src_merges:
        from openpyxl.utils import range_boundaries
        min_c, min_r, max_c, max_r = range_boundaries(m_str)
        new_min_r = DST_DATA_ROW if min_r == SRC_DATA_ROW else DST_KEG_ROW
        new_max_r = DST_DATA_ROW if max_r == SRC_DATA_ROW else DST_KEG_ROW
        new_range = f"{get_column_letter(min_c)}{new_min_r}:{get_column_letter(max_c)}{new_max_r}"
        # skip if already exists
        if new_range in [str(m) for m in ws.merged_cells.ranges]:
            continue
        try:
            ws.merge_cells(new_range)
        except Exception:
            pass

    # 5) Vertical merges (A/C/D/E/CR span both rows of a lab pair) — handle explicitly
    #    Src A26:A27, C26:C27, D26:D27, E26:E27, CR26:CR27 → A28:A29 etc.
    vertical_cols = [1, 3, 4, 5]  # A, C, D, E
    # find CR = last column
    for col in vertical_cols + [max_col]:
        col_letter = get_column_letter(col)
        rng = f"{col_letter}{DST_DATA_ROW}:{col_letter}{DST_KEG_ROW}"
        if rng not in [str(m) for m in ws.merged_cells.ranges]:
            try:
                ws.merge_cells(rng)
            except Exception:
                pass

    return f"OK (row 28 = {LAB_NAME})"


def main():
    print(f"Master: {MASTER}")
    wb = load_workbook(MASTER)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        msg = process_sheet(ws)
        print(f"  {sheet_name:12s} → {msg}")
    wb.save(MASTER)
    print("Saved.")

if __name__ == "__main__":
    main()
