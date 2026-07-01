"""Build REKAP UTILISASI SPP v3 tabel.xlsx — May–Dec 2026 with 5 removed sarana.

Removed from Mei 2026 onward:
  BRIEFING ROOM, LIQUID CARGO HANDLING SIMULATOR (LCHS),
  NAVIGATION AIDS SIMULATOR, SHIP OPERATIONAL LABORATORY (SOL),
  ENGINE ROOM GRAPHICS LABORATORY (ERGL).

Source layout cloned from FORMAT sheet of v2; only the 5 lab rows are deleted
(openpyxl delete_rows keeps merges/styles intact for surrounding rows).
"""
from copy import copy
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
SRC = HERE.parent / "master" / "REKAP UTILISASI SPP v2 tabel.xlsx"
DST = HERE.parent / "master" / "REKAP UTILISASI SPP v3 tabel.xlsx"

MONTHS_ID = {
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER",
}
YEAR = 2026

# Lab rows in FORMAT (row 12..24); rows to drop (5 removed sarana):
REMOVE_LAB_ROWS = [16, 17, 18, 19, 22]  # BRF, LCHS, NAS, SOL, ERGL

# Remaining 8 labs after removal, in original order:
KEEP_LAB_NAMES = [
    "CARGO HANDLING LABORATORY (CHL)",
    "ELECTRIC AND ELECTRONIC LABORATORY (EEL)",
    "ENGINEERING WORKSHOP",
    "MARINE ENGINEERING LABORATORY (MEL)",
    "COMPUTER BASED TRAINING 1&2",
    "ENGINE ROOM CERTIFICATION SIMULATOR (ERCS)",
    "LANGUAGE TRAINING LABORATORY ",
    "AUTOMATIC CONTROL SYSTEM LABORATORY (ACSL)",
]


def build():
    wb = load_workbook(SRC)
    template = wb["FORMAT"]

    # Drop every non-template sheet first.
    for name in list(wb.sheetnames):
        if name != "FORMAT":
            del wb[name]

    # Create month sheets by copying FORMAT.
    for month_num, month_name in MONTHS_ID.items():
        ws = wb.copy_worksheet(template)
        ws.title = month_name

        # Period header at row 7 col C: ": <Bulan> <Tahun>"
        ws.cell(row=7, column=3, value=f": {month_name.title()} {YEAR}")

        # Delete the 5 removed lab rows (bottom-up so indices stay valid).
        for r in sorted(REMOVE_LAB_ROWS, reverse=True):
            ws.delete_rows(r, 1)

        # Renumber NO column (A) for the remaining 8 labs (rows 12..19).
        for i, _ in enumerate(KEEP_LAB_NAMES, start=1):
            ws.cell(row=11 + i, column=1, value=i)

    # Remove the template last.
    del wb["FORMAT"]

    wb.save(DST)
    print(f"Wrote {DST}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    build()
