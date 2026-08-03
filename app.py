import os
import sqlite3
import secrets
import ipaddress
from calendar import monthrange, weekday
from pathlib import Path
from datetime import datetime, date, timezone, timedelta
from io import BytesIO

from flask import (
    Flask, render_template, request, jsonify, redirect, url_for,
    flash, send_from_directory, abort, send_file
)
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user, login_required, current_user
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

import json as _json
from db import init_db, get_conn, DATA_DIR, DB_PATH
from labs import (
    LABS,
    MONTHS_ID,
    MONTHS_SHORT_ID,
    LAB_DESCRIPTIONS,
    ABBREV_HELP,
    RETIRED_LABS,
    retired_in_period,
)
from excel_import import parse_master_format, parse_all_sheets
import holidays as holidays_mod
import backup as backup_mod

WIB = timezone(timedelta(hours=7))  # Asia/Jakarta, fixed offset (no DST)


def today_wib() -> date:
    return datetime.now(WIB).date()

UPLOAD_DIR = DATA_DIR / "uploads" / "photos"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_IMG = {"png", "jpg", "jpeg", "webp"}
ALLOWED_XLSX = {"xlsx", "xlsm"}

IG_POSTS_FILE = DATA_DIR / "instagram_posts.json"
IG_POSTS_BPSDM_FILE = DATA_DIR / "instagram_posts_bpsdm.json"

def _load_ig_posts(path=IG_POSTS_FILE) -> list[str]:
    """Load Instagram post URLs, filtering out placeholders."""
    if not path.exists():
        return []
    try:
        data = _json.loads(path.read_text("utf-8"))
        return [u for u in data.get("posts", [])
                if u and "PLACEHOLDER" not in u and "PASTE_" not in u and ("/p/" in u or "/reel/" in u)]
    except Exception:
        return []
MAX_IMAGE_BYTES = 2 * 1024 * 1024     # 2 MB per foto kegiatan
MAX_XLSX_BYTES  = 25 * 1024 * 1024    # 25 MB untuk file Excel Master Format
# PIL formats that map to the allowed extensions above
ALLOWED_PIL_FORMATS = {"PNG", "JPEG", "WEBP"}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SPP_SECRET", secrets.token_hex(16))
# Outer envelope: leaves room for Excel + form fields; image-specific cap enforced below.
app.config["MAX_CONTENT_LENGTH"] = MAX_XLSX_BYTES + 1024 * 1024

login_manager = LoginManager(app)
login_manager.login_view = "login"


class Admin(UserMixin):
    def __init__(self, row):
        self.id = str(row["id"])
        self.username = row["username"]


# ----- LAN-only write guard --------------------------------------------------
# Only LAN / loopback IPs may hit write endpoints. Public traffic via Cloudflare
# Tunnel, ngrok, etc. carries the real client IP in CF-Connecting-IP /
# X-Forwarded-For — those will be public addresses and get blocked here.
LAN_NETWORKS = [
    ipaddress.ip_network(n) for n in (
        "127.0.0.0/8", "10.0.0.0/8", "172.16.0.0/12", "192.168.0.0/16",
        "169.254.0.0/16", "::1/128", "fc00::/7", "fe80::/10",
    )
]
WRITE_METHODS = {"POST", "PUT", "PATCH", "DELETE"}


def _client_ip() -> str:
    raw = (
        request.headers.get("CF-Connecting-IP")
        or request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
        or request.remote_addr
        or ""
    )
    return raw


def _is_lan(ip: str) -> bool:
    try:
        addr = ipaddress.ip_address(ip)
    except ValueError:
        return False
    return any(addr in net for net in LAN_NETWORKS)


@app.before_request
def lan_only_writes():
    # Public-mode opt-out: SPP_ALLOW_PUBLIC_WRITES=1 disables the guard
    if os.environ.get("SPP_ALLOW_PUBLIC_WRITES") == "1":
        return None
    method_is_write = request.method in WRITE_METHODS
    path_is_write   = request.path.startswith("/admin")
    if not (method_is_write or path_is_write):
        return None
    ip = _client_ip()
    if _is_lan(ip):
        return None
    # Block — public viewers shouldn't be able to log in, edit, or upload either.
    if request.headers.get("Accept", "").startswith("application/json"):
        return jsonify({
            "ok": False, "code": "lan_only",
            "message": "Operasi tulis hanya dapat dilakukan dari jaringan lokal.",
        }), 403
    return (
        "<h3>Akses ditolak</h3>"
        "<p>Halaman ini hanya tersedia untuk pengguna di jaringan lokal STIP. "
        "Silakan akses dashboard versi publik (hanya lihat) di alamat utama.</p>"
    ), 403


@app.errorhandler(413)
def too_large(_e):
    flash(
        f"Berkas terlalu besar (batas {MAX_XLSX_BYTES // (1024*1024)} MB total). "
        f"Untuk foto kegiatan batasnya {MAX_IMAGE_BYTES // (1024*1024)} MB per foto.",
        "error",
    )
    return redirect(request.referrer or url_for("admin")), 303


@login_manager.user_loader
def load_user(uid):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM admin_user WHERE id = ?", (uid,)).fetchone()
        return Admin(row) if row else None


LOGO_PATH = Path(__file__).parent / "static" / "img" / "stip-logo.png"
STATIC_EXPORT = os.environ.get("SPP_STATIC_EXPORT") == "1"


@app.context_processor
def inject_clock_context():
    t = today_wib()
    return {
        "today_iso": t.isoformat(),
        "today_year": t.year,
        "holiday_today": holidays_mod.get_today(t),
        "lab_descriptions": LAB_DESCRIPTIONS,
        "abbrev_help": ABBREV_HELP,
        "months_short": MONTHS_SHORT_ID,
        "has_logo": LOGO_PATH.exists(),
        "static_export": STATIC_EXPORT,
        "retired_lab_codes": set(RETIRED_LABS.keys()),
        "retired_lab_meta": RETIRED_LABS,
    }


# ---------------- helpers ----------------

def all_labs():
    with get_conn() as conn:
        return [dict(r) for r in conn.execute("SELECT * FROM lab ORDER BY id").fetchall()]


def _filter_retired(labs_list, year, month):
    """Drop labs whose code is in RETIRED_LABS with retirement <= (year, month)."""
    out = []
    for lab in labs_list:
        code = lab.get("code") if isinstance(lab, dict) else None
        retired = RETIRED_LABS.get(code) if code else None
        if retired and (year, month) >= retired:
            continue
        out.append(lab)
    return out


def _filter_retired_year(labs_list, year):
    """For year view: zero out month cells for retired periods so totals don't lie."""
    out = []
    for lab in labs_list:
        code = lab.get("code") if isinstance(lab, dict) else None
        retired = RETIRED_LABS.get(code) if code else None
        if not retired:
            out.append(lab)
            continue
        ry, rm = retired
        if year < ry:
            out.append(lab)
            continue
        if year > ry:
            # Fully retired this whole year — drop.
            continue
        # Partial year: keep, but zero retired months.
        for m in range(rm, 13):
            cell = lab["months"].get(m)
            if cell:
                cell["fr"] = 0
                cell["jlh"] = 0
                cell["drs"] = 0.0
                cell["photos"] = 0
                cell["retired"] = True
        lab["fr_total"] = sum(v["fr"] for v in lab["months"].values())
        lab["jp_total"] = sum(v["jlh"] for v in lab["months"].values())
        lab["drs_total"] = sum(v["drs"] for v in lab["months"].values())
        lab["active_months"] = sum(1 for v in lab["months"].values() if v["drs"] > 0)
        lab["retired_from"] = (ry, rm)
        out.append(lab)
    return out


def period_summary(year, month):
    """Return per-lab totals + per-day series + day metadata + photos map + details map."""
    days = monthrange(year, month)[1]
    labs = all_labs()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT lab_id, day, fr, jlh, drs FROM entry WHERE year=? AND month=?",
            (year, month),
        ).fetchall()
        ket_rows = conn.execute(
            "SELECT lab_id, note FROM keterangan WHERE year=? AND month=?",
            (year, month),
        ).fetchall()
        photo_rows = conn.execute(
            "SELECT id, filename, event_date, lab_id, caption FROM photo "
            "WHERE event_date LIKE ? ORDER BY id",
            (f"{year}-{month:02d}-%",),
        ).fetchall()
        detail_rows = conn.execute(
            "SELECT lab_id, day, users, jabatan, activity FROM detail "
            "WHERE year=? AND month=? ORDER BY lab_id, day, id",
            (year, month),
        ).fetchall()

    by_lab = {lab["id"]: {
        "id": lab["id"], "code": lab["code"], "name": lab["name"],
        "fr_total": 0, "jp_total": 0, "drs_total": 0.0,
        "days": {d: {"fr": 0, "jlh": 0, "drs": 0.0} for d in range(1, days + 1)},
        "keterangan": "",
    } for lab in labs}

    for r in rows:
        if r["lab_id"] not in by_lab:
            continue
        cell = by_lab[r["lab_id"]]
        cell["days"][r["day"]] = {"fr": r["fr"], "jlh": r["jlh"], "drs": r["drs"]}
        cell["fr_total"] += r["fr"]
        cell["jp_total"] += r["jlh"]
        cell["drs_total"] += r["drs"]

    for k in ket_rows:
        if k["lab_id"] in by_lab:
            by_lab[k["lab_id"]]["keterangan"] = k["note"]

    holiday_map = holidays_mod.get_for_period(year, month)
    day_meta = {}
    for d in range(1, days + 1):
        dow = weekday(year, month, d)  # 0=Mon ... 6=Sun
        h = holiday_map.get(d)
        day_meta[d] = {
            "dow": dow,
            "is_weekend": dow >= 5,
            "holiday": h["name"] if h else None,
            "is_national": bool(h and h["is_national"]),
        }

    photos_map = {}
    for p in photo_rows:
        if not p["lab_id"]:
            continue
        try:
            d = int(p["event_date"].split("-")[2])
        except (ValueError, IndexError):
            continue
        photos_map.setdefault(p["lab_id"], {}).setdefault(d, []).append({
            "id": p["id"], "filename": p["filename"], "caption": p["caption"] or "",
        })

    details_map: dict[int, dict[int, list[dict]]] = {}
    for dr in detail_rows:
        details_map.setdefault(dr["lab_id"], {}).setdefault(dr["day"], []).append({
            "users": dr["users"] or "",
            "jabatan": dr["jabatan"] or "",
            "activity": dr["activity"] or "",
        })

    return list(by_lab.values()), days, day_meta, photos_map, details_map


def available_periods():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT DISTINCT year, month FROM entry ORDER BY year DESC, month DESC"
        ).fetchall()
    return [(r["year"], r["month"]) for r in rows]


def available_years():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT DISTINCT year FROM entry ORDER BY year DESC"
        ).fetchall()
    return [r["year"] for r in rows]


def year_summary(year):
    """Return per-lab × per-month aggregates for a full calendar year."""
    labs = all_labs()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT lab_id, month, "
            "SUM(fr) fr, SUM(jlh) jlh, SUM(drs) drs "
            "FROM entry WHERE year=? GROUP BY lab_id, month",
            (year,),
        ).fetchall()
        ket_rows = conn.execute(
            "SELECT lab_id, month, note FROM keterangan WHERE year=?",
            (year,),
        ).fetchall()
        photo_rows = conn.execute(
            "SELECT lab_id, event_date FROM photo "
            "WHERE event_date LIKE ? AND lab_id IS NOT NULL",
            (f"{year}-%",),
        ).fetchall()

    by_lab = {lab["id"]: {
        "id": lab["id"], "code": lab["code"], "name": lab["name"],
        "fr_total": 0, "jp_total": 0, "drs_total": 0.0,
        "months": {m: {"fr": 0, "jlh": 0, "drs": 0.0, "photos": 0} for m in range(1, 13)},
        "active_months": 0,
        "keterangan": {},
    } for lab in labs}

    for r in rows:
        lab = by_lab.get(r["lab_id"])
        if not lab:
            continue
        lab["months"][r["month"]] = {
            "fr": r["fr"] or 0, "jlh": r["jlh"] or 0,
            "drs": float(r["drs"] or 0), "photos": 0,
        }
        lab["fr_total"] += r["fr"] or 0
        lab["jp_total"] += r["jlh"] or 0
        lab["drs_total"] += float(r["drs"] or 0)

    for k in ket_rows:
        if k["lab_id"] in by_lab:
            by_lab[k["lab_id"]]["keterangan"][k["month"]] = k["note"]

    for p in photo_rows:
        lab = by_lab.get(p["lab_id"])
        if not lab:
            continue
        try:
            m = int(p["event_date"].split("-")[1])
        except (ValueError, IndexError):
            continue
        if 1 <= m <= 12:
            lab["months"][m]["photos"] += 1

    for lab in by_lab.values():
        lab["active_months"] = sum(
            1 for v in lab["months"].values() if v["drs"] > 0
        )

    return list(by_lab.values())


def allowed(filename, allowed_set):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed_set


def _stream_size(stream) -> int:
    """Return the byte length of an upload stream by seeking to the end."""
    pos = stream.tell()
    stream.seek(0, os.SEEK_END)
    size = stream.tell()
    stream.seek(pos)
    return size


def validate_image_upload(file_storage) -> tuple[bool, str | None]:
    """Check extension, size cap, and that the file is a real image."""
    if not allowed(file_storage.filename, ALLOWED_IMG):
        return False, (
            "Format foto tidak didukung. Gunakan ekstensi umum: "
            ".png, .jpg, .jpeg, atau .webp."
        )
    size = _stream_size(file_storage.stream)
    if size > MAX_IMAGE_BYTES:
        mb = size / (1024 * 1024)
        return False, (
            f"Ukuran foto {mb:.2f} MB melebihi batas {MAX_IMAGE_BYTES // (1024*1024)} MB. "
            "Kompres atau ubah ukuran foto sebelum mengunggah."
        )
    if size == 0:
        return False, "Berkas foto kosong."
    try:
        from PIL import Image, UnidentifiedImageError
        try:
            file_storage.stream.seek(0)
            with Image.open(file_storage.stream) as img:
                img.verify()
                fmt = (img.format or "").upper()
        except (UnidentifiedImageError, OSError):
            return False, "Berkas tidak terdeteksi sebagai foto yang valid."
        finally:
            file_storage.stream.seek(0)
    except ImportError:
        # PIL not installed; fall back to extension-only check
        pass
    else:
        if fmt not in ALLOWED_PIL_FORMATS:
            return False, (
                f"Isi berkas adalah format {fmt or 'tidak dikenal'}, bukan foto biasa. "
                "Gunakan PNG, JPEG, atau WebP."
            )
    return True, None


# ---------------- public routes ----------------

@app.route("/")
def index():
    periods = available_periods()
    today = today_wib()
    if periods:
        year, month = periods[0]
    else:
        year, month = today.year, today.month
    try:
        year = int(request.args.get("year", year))
        month = int(request.args.get("month", month))
    except ValueError:
        pass
    summary, days, day_meta, photos_map, details_map = period_summary(year, month)
    summary = _filter_retired(summary, year, month)
    retired = retired_in_period(year, month)
    return render_template(
        "index.html",
        labs=summary, days=days, year=year, month=month,
        day_meta=day_meta, photos_map=photos_map, details_map=details_map,
        month_name=MONTHS_ID.get(month, ""),
        periods=periods, months=MONTHS_ID,
        retired_labs=retired,
        ig_posts=_load_ig_posts(),
        ig_posts_bpsdm=_load_ig_posts(IG_POSTS_BPSDM_FILE),
    )


@app.route("/api/activity-check")
@login_required
def api_activity_check():
    """Returns whether a lab has FR/JLH/DRS activity on a given date."""
    try:
        lab_id = int(request.args["lab_id"])
        dt = datetime.strptime(request.args["date"], "%Y-%m-%d").date()
    except (KeyError, ValueError):
        return jsonify({"ok": False, "error": "bad args"}), 400
    with get_conn() as conn:
        lab = conn.execute("SELECT name FROM lab WHERE id=?", (lab_id,)).fetchone()
        row = conn.execute(
            "SELECT fr, jlh, drs FROM entry "
            "WHERE lab_id=? AND year=? AND month=? AND day=?",
            (lab_id, dt.year, dt.month, dt.day),
        ).fetchone()
    has_activity = bool(row and (row["fr"] or row["jlh"] or row["drs"]))
    return jsonify({
        "ok": True,
        "lab_name": lab["name"] if lab else None,
        "has_activity": has_activity,
        "fr": (row["fr"] if row else 0),
        "jlh": (row["jlh"] if row else 0),
        "drs": (row["drs"] if row else 0),
    })


@app.route("/api/period")
def api_period():
    year = int(request.args["year"])
    month = int(request.args["month"])
    summary, days, day_meta, photos_map, details_map = period_summary(year, month)
    summary = _filter_retired(summary, year, month)
    return jsonify({
        "year": year, "month": month, "days": days,
        "day_meta": day_meta,
        "photos": photos_map,
        "details": details_map,
        "retired_labs": [{"code": c, "name": n} for c, n in retired_in_period(year, month)],
        "labs": [{
            "id": l["id"], "code": l["code"], "name": l["name"],
            "fr_total": l["fr_total"], "jp_total": l["jp_total"], "drs_total": l["drs_total"],
            "days": l["days"], "keterangan": l["keterangan"],
        } for l in summary],
    })


@app.route("/year")
def year_view():
    today = today_wib()
    years = available_years() or [today.year]
    try:
        year = int(request.args.get("year", years[0]))
    except ValueError:
        year = years[0]
    summary = year_summary(year)
    summary = _filter_retired_year(summary, year)
    return render_template(
        "year.html",
        labs=summary, year=year,
        months=MONTHS_ID, months_short=MONTHS_SHORT_ID,
        years=years,
        retired_labs=[{"code": c, "name": n, "from_month": RETIRED_LABS[c][1], "from_year": RETIRED_LABS[c][0]}
                      for c, n in retired_in_period(year, 12) if RETIRED_LABS[c][0] <= year],
    )


@app.route("/gallery")
def gallery():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT p.*, l.code AS lab_code, l.name AS lab_name "
            "FROM photo p LEFT JOIN lab l ON p.lab_id = l.id "
            "ORDER BY p.event_date DESC, p.id DESC"
        ).fetchall()
    return render_template("gallery.html", photos=rows, labs=all_labs())


@app.route("/uploads/photos/<path:filename>")
def serve_photo(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@app.route("/profile")
def profile_spp():
    return render_template("profile.html")


# ---------------- auth ----------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM admin_user WHERE username=?", (u,)).fetchone()
        if row and check_password_hash(row["pw_hash"], p):
            login_user(Admin(row))
            return redirect(request.args.get("next") or url_for("admin"))
        flash("Username atau password salah.", "error")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))


@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    cur = request.form.get("current", "")
    new = request.form.get("new", "")
    if len(new) < 4:
        flash("Password baru minimal 4 karakter.", "error")
        return redirect(url_for("admin"))
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM admin_user WHERE id=?", (current_user.id,)).fetchone()
        if not check_password_hash(row["pw_hash"], cur):
            flash("Password saat ini salah.", "error")
            return redirect(url_for("admin"))
        conn.execute("UPDATE admin_user SET pw_hash=? WHERE id=?",
                     (generate_password_hash(new), current_user.id))
        conn.commit()
    flash("Password berhasil diubah.", "success")
    return redirect(url_for("admin"))


# ---------------- admin routes ----------------

@app.route("/admin")
@login_required
def admin():
    today = today_wib()
    try:
        year = int(request.args.get("year", today.year))
        month = int(request.args.get("month", today.month))
    except ValueError:
        year, month = today.year, today.month
    summary, days, day_meta, _, _ = period_summary(year, month)
    summary = _filter_retired(summary, year, month)
    holiday_count = sum(1 for d in day_meta.values() if d["holiday"])
    return render_template(
        "admin.html",
        labs=summary, days=days, year=year, month=month,
        day_meta=day_meta, holiday_count=holiday_count,
        month_name=MONTHS_ID.get(month, ""), months=MONTHS_ID,
        periods=available_periods(),
        retired_labs=retired_in_period(year, month),
    )


@app.route("/admin/entry", methods=["POST"])
@login_required
def admin_save_entry():
    data = request.get_json(force=True)
    lab_id = int(data["lab_id"])
    year   = int(data["year"])
    month  = int(data["month"])
    day    = int(data["day"])
    fr     = int(data.get("fr") or 0)
    jlh    = int(data.get("jlh") or 0)
    drs    = float(data.get("drs") or 0)
    confirm_delete = bool(data.get("confirm_delete"))

    with get_conn() as conn:
        existing = conn.execute(
            "SELECT fr, jlh, drs FROM entry "
            "WHERE lab_id=? AND year=? AND month=? AND day=?",
            (lab_id, year, month, day),
        ).fetchone()
        clearing = (fr == 0 and jlh == 0 and drs == 0)
        had_data = bool(existing and (existing["fr"] or existing["jlh"] or existing["drs"]))

        # Refuse to silently destroy non-zero data without an explicit confirm flag.
        if clearing and had_data and not confirm_delete:
            return jsonify({
                "ok": False, "code": "confirm_required",
                "previous": {"fr": existing["fr"], "jlh": existing["jlh"], "drs": existing["drs"]},
                "message": "Sel ini berisi data. Konfirmasi penghapusan diperlukan.",
            }), 409

        if clearing:
            conn.execute(
                "DELETE FROM entry WHERE lab_id=? AND year=? AND month=? AND day=?",
                (lab_id, year, month, day),
            )
        else:
            conn.execute(
                "INSERT INTO entry (lab_id, year, month, day, fr, jlh, drs) VALUES (?,?,?,?,?,?,?) "
                "ON CONFLICT(lab_id, year, month, day) DO UPDATE SET "
                "fr=excluded.fr, jlh=excluded.jlh, drs=excluded.drs",
                (lab_id, year, month, day, fr, jlh, drs),
            )
        conn.commit()
    return jsonify({"ok": True})


@app.route("/admin/delete-period", methods=["POST"])
@login_required
def admin_delete_period():
    try:
        year = int(request.form["year"])
        month = int(request.form["month"])
    except (KeyError, ValueError):
        flash("Periode tidak valid.", "error")
        return redirect(url_for("admin"))

    expected = f"HAPUS {MONTHS_ID[month].upper()} {year}"
    typed = (request.form.get("confirm_phrase") or "").strip().upper()
    if typed != expected:
        flash(
            f"Konfirmasi tidak cocok. Ketik persis: \"{expected}\" untuk menghapus.",
            "error",
        )
        return redirect(url_for("admin", year=year, month=month))

    backup_mod.create_backup(f"pre_delete_{year}_{month:02d}")
    with get_conn() as conn:
        n_entry = conn.execute(
            "DELETE FROM entry WHERE year=? AND month=?", (year, month)
        ).rowcount
        n_ket = conn.execute(
            "DELETE FROM keterangan WHERE year=? AND month=?", (year, month)
        ).rowcount
        conn.commit()
    flash(
        f"Periode {MONTHS_ID[month]} {year} dihapus ({n_entry} entri, {n_ket} keterangan). "
        f"Snapshot tersimpan di data/backups/ — bisa dipulihkan dari menu Backup.",
        "success",
    )
    return redirect(url_for("admin"))


@app.route("/admin/backups")
@login_required
def admin_backups():
    return render_template("backups.html", backups=backup_mod.list_backups())


@app.route("/admin/backup-now", methods=["POST"])
@login_required
def admin_backup_now():
    path = backup_mod.create_backup("manual")
    if path:
        flash(f"Snapshot manual tersimpan: {path.name}", "success")
    else:
        flash("Tidak ada database untuk di-backup.", "error")
    return redirect(url_for("admin_backups"))


@app.route("/admin/backup-restore", methods=["POST"])
@login_required
def admin_backup_restore():
    name = (request.form.get("name") or "").strip()
    if (request.form.get("confirm_phrase") or "").strip().upper() != "PULIHKAN":
        flash("Konfirmasi tidak cocok. Ketik PULIHKAN untuk mengembalikan database.", "error")
        return redirect(url_for("admin_backups"))
    try:
        backup_mod.restore_backup(name)
    except FileNotFoundError:
        flash("Berkas backup tidak ditemukan.", "error")
        return redirect(url_for("admin_backups"))
    except ValueError:
        flash("Nama backup tidak valid.", "error")
        return redirect(url_for("admin_backups"))
    flash(
        f"Database dipulihkan dari {name}. Snapshot database sebelumnya juga tersimpan untuk berjaga-jaga.",
        "success",
    )
    return redirect(url_for("admin"))


@app.route("/admin/keterangan", methods=["POST"])
@login_required
def admin_save_keterangan():
    data = request.get_json(force=True)
    lab_id = int(data["lab_id"])
    year = int(data["year"])
    month = int(data["month"])
    note = (data.get("note") or "").strip()
    with get_conn() as conn:
        if not note:
            conn.execute(
                "DELETE FROM keterangan WHERE lab_id=? AND year=? AND month=?",
                (lab_id, year, month),
            )
        else:
            conn.execute(
                "INSERT INTO keterangan (lab_id, year, month, note) VALUES (?,?,?,?) "
                "ON CONFLICT(lab_id, year, month) DO UPDATE SET note=excluded.note",
                (lab_id, year, month, note),
            )
        conn.commit()
    return jsonify({"ok": True})


@app.route("/admin/import", methods=["POST"])
@login_required
def admin_import():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Pilih file Excel terlebih dahulu.", "error")
        return redirect(url_for("admin"))
    if not allowed(f.filename, ALLOWED_XLSX):
        flash("Hanya file .xlsx atau .xlsm yang didukung.", "error")
        return redirect(url_for("admin"))
    overwrite = request.form.get("overwrite") == "1"
    # Server-side guardrail: destructive overwrite needs typed confirmation.
    if overwrite and request.form.get("overwrite_confirm", "").strip().upper() != "OVERWRITE":
        flash(
            "Mode 'Timpa data' membutuhkan konfirmasi. Ketik OVERWRITE pada kolom konfirmasi "
            "untuk melanjutkan, atau hilangkan centang untuk mode aman (tidak menimpa data).",
            "error",
        )
        return redirect(url_for("admin"))

    try:
        tmp = DATA_DIR / "_import_tmp.xlsx"
        f.save(tmp)
        sheet_results = parse_all_sheets(tmp)
    except Exception as e:
        flash(f"Gagal membaca Excel: {e}", "error")
        return redirect(url_for("admin"))

    parseable = [s for s in sheet_results if s["ok"]]
    if not parseable:
        skipped_lines = "; ".join(f"{s['sheet']}: {s.get('error','-')}" for s in sheet_results)
        flash(
            "Tidak ada worksheet yang sesuai format Master. Periksa header BULAN di tiap sheet. "
            f"Detail: {skipped_lines}",
            "error",
        )
        return redirect(url_for("admin"))

    backup_mod.create_backup(
        f"pre_import_{len(parseable)}sheets_{'overwrite' if overwrite else 'merge'}"
    )

    per_period_stats = []
    with get_conn() as conn:
        for parsed in parseable:
            year, month = parsed["year"], parsed["month"]
            inserted = updated = skipped = 0
            details_added = 0
            if overwrite:
                conn.execute("DELETE FROM entry WHERE year=? AND month=?", (year, month))
                conn.execute("DELETE FROM keterangan WHERE year=? AND month=?", (year, month))

            # Details are sourced exclusively from the Excel KETERANGAN column,
            # so each re-import is idempotent: wipe the labs this sheet touches,
            # then re-insert. This intentionally bypasses safe/overwrite mode —
            # the master file is the source of truth for detail rows.
            touched_labs = {d["lab_id"] for d in parsed.get("details", [])}
            touched_labs.update(parsed["keterangan"].keys())
            for lab_id in touched_labs:
                conn.execute(
                    "DELETE FROM detail WHERE lab_id=? AND year=? AND month=?",
                    (lab_id, year, month),
                )
            for d in parsed.get("details", []):
                conn.execute(
                    "INSERT INTO detail (lab_id, year, month, day, users, jabatan, activity) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (d["lab_id"], year, month, d["day"],
                     d.get("users", ""), d.get("jabatan", ""), d.get("activity", "")),
                )
                details_added += 1

            for r in parsed["rows"]:
                if overwrite:
                    conn.execute(
                        "INSERT INTO entry (lab_id, year, month, day, fr, jlh, drs) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (r["lab_id"], year, month, r["day"], r["fr"], r["jlh"], r["drs"]),
                    )
                    inserted += 1
                    continue
                existing = conn.execute(
                    "SELECT fr, jlh, drs FROM entry "
                    "WHERE lab_id=? AND year=? AND month=? AND day=?",
                    (r["lab_id"], year, month, r["day"]),
                ).fetchone()
                if existing is None:
                    conn.execute(
                        "INSERT INTO entry (lab_id, year, month, day, fr, jlh, drs) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (r["lab_id"], year, month, r["day"], r["fr"], r["jlh"], r["drs"]),
                    )
                    inserted += 1
                elif existing["fr"] == 0 and existing["jlh"] == 0 and existing["drs"] == 0:
                    conn.execute(
                        "UPDATE entry SET fr=?, jlh=?, drs=? "
                        "WHERE lab_id=? AND year=? AND month=? AND day=?",
                        (r["fr"], r["jlh"], r["drs"], r["lab_id"], year, month, r["day"]),
                    )
                    updated += 1
                else:
                    skipped += 1

            for lab_id, note in parsed["keterangan"].items():
                existing_note = conn.execute(
                    "SELECT note FROM keterangan WHERE lab_id=? AND year=? AND month=?",
                    (lab_id, year, month),
                ).fetchone()
                if existing_note is None or overwrite or not (existing_note["note"] or "").strip():
                    conn.execute(
                        "INSERT INTO keterangan (lab_id, year, month, note) VALUES (?,?,?,?) "
                        "ON CONFLICT(lab_id, year, month) DO UPDATE SET note=excluded.note",
                        (lab_id, year, month, note),
                    )

            per_period_stats.append({
                "sheet": parsed["sheet"], "year": year, "month": month,
                "inserted": inserted, "updated": updated, "skipped": skipped,
                "details": details_added,
            })
        conn.commit()

    summary_parts = []
    total_in = total_up = total_sk = total_det = 0
    for s in per_period_stats:
        label = f"{MONTHS_ID[s['month']]} {s['year']}"
        bits = []
        if s["inserted"]: bits.append(f"{s['inserted']} entri baru")
        if s["updated"]:  bits.append(f"{s['updated']} entri kosong diisi")
        if s["skipped"]:  bits.append(f"{s['skipped']} DILEWATI")
        if s["details"]:  bits.append(f"{s['details']} detail kegiatan")
        summary_parts.append(f"{label} ({', '.join(bits) if bits else 'tidak ada perubahan'})")
        total_in += s["inserted"]; total_up += s["updated"]; total_sk += s["skipped"]
        total_det += s["details"]

    skipped_sheets = [s for s in sheet_results if not s["ok"]]
    msg_head = (
        f"Import berhasil &mdash; {len(per_period_stats)} bulan diproses "
        f"({total_in} baru, {total_up} diisi"
    )
    if not overwrite and total_sk:
        msg_head += f", {total_sk} DILEWATI karena sudah berisi data"
    if total_det:
        msg_head += f", {total_det} baris detail kegiatan dari kolom KETERANGAN"
    msg_head += "). " + " | ".join(summary_parts)
    if skipped_sheets:
        msg_head += (
            " &mdash; Sheet tanpa data Master: "
            + ", ".join(f"\"{s['sheet']}\"" for s in skipped_sheets)
        )
    msg_head += ". Snapshot database tersimpan di data/backups/."
    flash(msg_head, "success")
    target = per_period_stats[0] if per_period_stats else None
    if target:
        return redirect(url_for("admin", year=target["year"], month=target["month"]))
    return redirect(url_for("admin"))


@app.route("/admin/photo", methods=["POST"])
@login_required
def admin_upload_photo():
    f = request.files.get("photo")
    event_date = request.form.get("event_date", "").strip()
    lab_id_raw = (request.form.get("lab_id") or "").strip()
    caption = request.form.get("caption", "").strip()

    if not f or not f.filename:
        flash("Pilih foto yang akan diunggah.", "error")
        return redirect(url_for("admin"))
    ok, err = validate_image_upload(f)
    if not ok:
        flash(err, "error")
        return redirect(url_for("admin"))
    try:
        dt = datetime.strptime(event_date, "%Y-%m-%d").date()
    except ValueError:
        flash("Tanggal kegiatan wajib diisi (format YYYY-MM-DD).", "error")
        return redirect(url_for("admin"))

    if dt > today_wib():
        flash("Tanggal kegiatan tidak boleh di masa depan.", "error")
        return redirect(url_for("admin"))

    lab_id = int(lab_id_raw) if lab_id_raw else None

    # Activity guard: a lab-specific photo requires recorded activity on that date.
    if lab_id is not None:
        with get_conn() as conn:
            lab_row = conn.execute("SELECT name FROM lab WHERE id=?", (lab_id,)).fetchone()
            act = conn.execute(
                "SELECT fr, jlh, drs FROM entry "
                "WHERE lab_id=? AND year=? AND month=? AND day=?",
                (lab_id, dt.year, dt.month, dt.day),
            ).fetchone()
        if not act or (act["fr"] == 0 and act["jlh"] == 0 and act["drs"] == 0):
            lab_name = lab_row["name"] if lab_row else f"lab #{lab_id}"
            flash(
                f"Tidak dapat mengunggah foto: belum ada aktivitas tercatat untuk "
                f"{lab_name} pada tanggal {event_date}. "
                f"Catat data utilisasi (FR/JLH/DRS) terlebih dahulu, atau pilih "
                f"\"Umum / tidak spesifik\" untuk foto yang tidak terkait lab tertentu.",
                "error",
            )
            return redirect(url_for("admin"))

    ext = f.filename.rsplit(".", 1)[1].lower()
    safe = secure_filename(f.filename.rsplit(".", 1)[0])[:40] or "photo"
    new_name = f"{event_date}_{safe}_{secrets.token_hex(4)}.{ext}"
    f.save(UPLOAD_DIR / new_name)

    with get_conn() as conn:
        conn.execute(
            "INSERT INTO photo (filename, event_date, lab_id, caption) VALUES (?,?,?,?)",
            (new_name, event_date, lab_id, caption),
        )
        conn.commit()
    flash("Foto berhasil diunggah.", "success")
    return redirect(url_for("admin"))


@app.route("/admin/sync-holidays", methods=["POST"])
@login_required
def admin_sync_holidays():
    try:
        year = int(request.form.get("year", today_wib().year))
    except ValueError:
        year = today_wib().year
    result = holidays_mod.sync_year(year)
    if result["ok"]:
        flash(
            f"Hari libur {year} berhasil disinkronkan: {result['count']} entri "
            f"(sumber: {result['source']}).",
            "success",
        )
    else:
        flash(
            f"Gagal sinkronisasi hari libur {year}: {result['error']}. "
            f"Pastikan PC terhubung internet, lalu coba lagi.",
            "error",
        )
    return redirect(request.referrer or url_for("admin"))


@app.route("/admin/photo/<int:pid>/edit", methods=["POST"])
@login_required
def admin_edit_photo(pid):
    with get_conn() as conn:
        existing = conn.execute(
            "SELECT id, filename, event_date, lab_id, caption FROM photo WHERE id=?",
            (pid,),
        ).fetchone()
    if not existing:
        abort(404)

    event_date = (request.form.get("event_date") or "").strip()
    lab_id_raw = (request.form.get("lab_id") or "").strip()
    caption = (request.form.get("caption") or "").strip()
    new_file = request.files.get("photo")
    redirect_to = request.referrer or url_for("gallery")

    try:
        dt = datetime.strptime(event_date, "%Y-%m-%d").date()
    except ValueError:
        flash("Tanggal kegiatan wajib diisi (format YYYY-MM-DD).", "error")
        return redirect(redirect_to)

    if dt > today_wib():
        flash("Tanggal kegiatan tidak boleh di masa depan.", "error")
        return redirect(redirect_to)

    lab_id = int(lab_id_raw) if lab_id_raw else None

    # Activity guard mirrors the upload route: a lab-specific photo needs
    # recorded utilisation on that date. "Umum / tidak spesifik" bypasses it.
    if lab_id is not None:
        with get_conn() as conn:
            lab_row = conn.execute("SELECT name FROM lab WHERE id=?", (lab_id,)).fetchone()
            act = conn.execute(
                "SELECT fr, jlh, drs FROM entry "
                "WHERE lab_id=? AND year=? AND month=? AND day=?",
                (lab_id, dt.year, dt.month, dt.day),
            ).fetchone()
        if not act or (act["fr"] == 0 and act["jlh"] == 0 and act["drs"] == 0):
            lab_name = lab_row["name"] if lab_row else f"lab #{lab_id}"
            flash(
                f"Tidak dapat menyimpan: belum ada aktivitas tercatat untuk "
                f"{lab_name} pada tanggal {event_date}. "
                f"Catat data utilisasi (FR/JLH/DRS) dulu, atau pilih "
                f"\"Umum / tidak spesifik\".",
                "error",
            )
            return redirect(redirect_to)

    replacement_name: str | None = None
    if new_file and new_file.filename:
        ok, err = validate_image_upload(new_file)
        if not ok:
            flash(err, "error")
            return redirect(redirect_to)
        ext = new_file.filename.rsplit(".", 1)[1].lower()
        safe = secure_filename(new_file.filename.rsplit(".", 1)[0])[:40] or "photo"
        replacement_name = f"{event_date}_{safe}_{secrets.token_hex(4)}.{ext}"
        new_file.save(UPLOAD_DIR / replacement_name)

    final_name = replacement_name or existing["filename"]
    with get_conn() as conn:
        conn.execute(
            "UPDATE photo SET filename=?, event_date=?, lab_id=?, caption=? WHERE id=?",
            (final_name, event_date, lab_id, caption, pid),
        )
        conn.commit()

    # Only unlink the old file *after* the DB row has been updated so a crash
    # mid-update can never leave a row pointing at a missing file.
    if replacement_name and replacement_name != existing["filename"]:
        try:
            (UPLOAD_DIR / existing["filename"]).unlink(missing_ok=True)
        except OSError:
            pass

    flash("Foto diperbarui.", "success")
    return redirect(redirect_to)


@app.route("/admin/photo/<int:pid>/delete", methods=["POST"])
@login_required
def admin_delete_photo(pid):
    with get_conn() as conn:
        row = conn.execute("SELECT filename FROM photo WHERE id=?", (pid,)).fetchone()
        if not row:
            abort(404)
        try:
            (UPLOAD_DIR / row["filename"]).unlink(missing_ok=True)
        except OSError:
            pass
        conn.execute("DELETE FROM photo WHERE id=?", (pid,))
        conn.commit()
    flash("Foto dihapus.", "success")
    return redirect(request.referrer or url_for("admin"))


# ---------------- Excel export ----------------

@app.route("/export")
def export_excel():
    try:
        year = int(request.args["year"])
        month = int(request.args["month"])
    except (KeyError, ValueError):
        abort(400)
    summary, days, _day_meta, _photos, _details = period_summary(year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MASTER FORMAT"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="E8EEF7")

    ws["B2"] = "Rekapitulasi Utilisasi Sarana Praktek Pelaut"
    ws["B2"].font = Font(bold=True, size=12)
    ws["B3"] = f"BULAN : {MONTHS_ID[month]} {year}"

    # Header row 5-7 (NO, NAMA LAB, JF, JP, DURASI, then days)
    headers_top = ["NO", "NAMA LAB", "JF", "JP", "DURASI (JAM)", "TANGGAL HARI KERJA"]
    for i, h in enumerate(headers_top, start=1):
        ws.cell(row=5, column=i, value=h).font = bold
        ws.cell(row=5, column=i).alignment = center
        ws.cell(row=5, column=i).fill = fill
    ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=5 + days * 3)
    ws.cell(row=5, column=6 + days * 3, value="KETERANGAN").font = bold

    for d in range(1, days + 1):
        col = 6 + (d - 1) * 3
        ws.cell(row=6, column=col, value=d).font = bold
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 2)
        for j, sub in enumerate(["FR", "JLH", "DRS"]):
            c = ws.cell(row=7, column=col + j, value=sub)
            c.font = bold
            c.alignment = center
            c.fill = fill

    # Body
    for i, lab in enumerate(summary, start=1):
        r = 7 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=lab["name"])
        ws.cell(row=r, column=3, value=lab["fr_total"])
        ws.cell(row=r, column=4, value=lab["jp_total"])
        ws.cell(row=r, column=5, value=round(lab["drs_total"], 2))
        for d in range(1, days + 1):
            cell = lab["days"][d]
            base = 6 + (d - 1) * 3
            if cell["fr"]:  ws.cell(row=r, column=base,     value=cell["fr"])
            if cell["jlh"]: ws.cell(row=r, column=base + 1, value=cell["jlh"])
            if cell["drs"]: ws.cell(row=r, column=base + 2, value=cell["drs"])
        ws.cell(row=r, column=6 + days * 3, value=lab["keterangan"])

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    for cidx in range(3, 6 + days * 3 + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(cidx)].width = 6

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"Rekap_Utilisasi_{MONTHS_ID[month]}_{year}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname,
    )


# ---------------- entrypoint ----------------

def parse_args():
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--host", default=os.environ.get("SPP_HOST", "0.0.0.0"),
                   help="127.0.0.1 = PC ini saja, 0.0.0.0 = jaringan lokal (default)")
    p.add_argument("--port", type=int, default=int(os.environ.get("SPP_PORT", "5000")))
    p.add_argument("--debug", action="store_true")
    return p.parse_args()


if __name__ == "__main__":
    init_db(
        default_admin_user=os.environ.get("SPP_ADMIN_USER", "admin"),
        default_admin_pw=os.environ.get("SPP_ADMIN_PW", "admin"),
    )
    # Best-effort holiday prefetch (silent if offline)
    cur_year = today_wib().year
    for y in (cur_year, cur_year + 1):
        if not holidays_mod.has_any_for_year(y):
            try:
                r = holidays_mod.sync_year(y)
                if r["ok"]:
                    print(f"  Hari libur {y}: {r['count']} entri disinkronkan dari {r['source']}.")
                else:
                    print(f"  (peringatan) Sinkronisasi hari libur {y} dilewati: {r['error']}")
            except Exception as e:
                print(f"  (peringatan) Sinkronisasi hari libur {y} gagal: {e}")
    args = parse_args()
    print(f"\n  Dashboard Utilisasi SPP siap di: http://{args.host}:{args.port}\n")
    app.run(host=args.host, port=args.port, debug=args.debug)
